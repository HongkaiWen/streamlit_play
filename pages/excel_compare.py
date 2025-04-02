import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import re

from access_log import log_access


def remove_html_tags(text):
    """去除 HTML 标签的函数"""
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text)


def compare_excels(file1, file2, primary_keys, selected_columns):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 只保留共有的列且排除主键
    common_columns = [col for col in selected_columns if
                      col in df1.columns and col in df2.columns and col not in primary_keys]

    # 只列出两个 Excel 主键可关联到的行
    merged_df = pd.merge(df1, df2, on=primary_keys, how='inner', suffixes=('_A', '_B'))

    # 对比结果
    comparison_result = []
    for _, row in merged_df.iterrows():
        is_match = True
        for col in common_columns:
            col_a = f'{col}_A'
            col_b = f'{col}_B'
            if row[col_a] != row[col_b]:
                is_match = False
                break
        comparison_result.append('Y' if is_match else 'N')

    merged_df['比对结果'] = comparison_result

    # 标记不一致的列
    for col in common_columns:
        col_a = f'{col}_A'
        col_b = f'{col}_B'
        merged_df[f'{col}_A'] = [
            f'<span style="color:red">{val}</span>' if merged_df['比对结果'][i] == 'N' and val != merged_df[col_b][
                i] else val for i, val in
            enumerate(merged_df[col_a])]
        merged_df[f'{col}_B'] = [
            f'<span style="color:red">{val}</span>' if merged_df['比对结果'][i] == 'N' and val != merged_df[col_a][
                i] else val for i, val in
            enumerate(merged_df[col_b])]

    # 第一个表格有第二个表格没有的数据
    df1_only = df1[~df1[primary_keys].apply(tuple, axis=1).isin(df2[primary_keys].apply(tuple, axis=1))]
    # 第二个表格有第一个表格没有的数据
    df2_only = df2[~df2[primary_keys].apply(tuple, axis=1).isin(df1[primary_keys].apply(tuple, axis=1))]

    return merged_df, df1_only, df2_only


def highlight_excel(workbook, df, primary_keys, common_columns):
    sheet = workbook.active
    header = next(dataframe_to_rows(df, index=False, header=True))
    sheet.append(header)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            col_name = df.columns[c_idx - 1]
            base_col = col_name.replace("_A", "").replace("_B", "")
            if base_col in common_columns:
                col_a = f'{base_col}_A'
                col_b = f'{base_col}_B'
                if df['比对结果'].iloc[r_idx - 2] == 'N' and df[col_a].iloc[r_idx - 2] != df[col_b].iloc[r_idx - 2]:
                    sheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                           fill_type='solid')
            # 去除 HTML 标签
            clean_value = remove_html_tags(str(value))
            sheet.cell(row=r_idx, column=c_idx).value = clean_value
    return workbook


def write_df_to_excel(workbook, df, sheet_name):
    sheet = workbook.create_sheet(sheet_name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx).value = value
    return workbook


def check_primary_key_uniqueness(df, primary_keys):
    """检查主键的唯一性"""
    grouped = df.groupby(primary_keys).size()
    if any(grouped > 1):
        return False
    return True


def excel_compare():
    log_access("excel_compare")
    st.title("Excel 文件对比工具")

    # 使用侧边栏进行文件上传和参数选择
    with st.sidebar:
        st.subheader("文件上传")
        st.write("请上传两个 Excel 文件进行对比。")
        file1 = st.file_uploader("上传第一个 Excel 文件", type=["xlsx", "xls"])
        file2 = st.file_uploader("上传第二个 Excel 文件", type=["xlsx", "xls"])

        if file1 and file2:
            df1 = pd.read_excel(file1)
            df2 = pd.read_excel(file2)

            # 获取所有列名
            all_columns = list(set(df1.columns).intersection(set(df2.columns)))

            st.subheader("参数设置")
            st.write("请选择用于对比的主键和参与对比的列。")
            # 选择主键
            primary_keys = st.multiselect("选择主键", all_columns)

            # 选择参与对比的列
            available_columns = [col for col in all_columns if col not in primary_keys]
            selected_columns = st.multiselect("选择参与对比的列", available_columns, default=available_columns)

            compare_button = st.button("开始对比", use_container_width=True)

    if 'compare_button' in locals() and compare_button:
        if not primary_keys:
            st.error("请至少选择一个主键。")
        else:
            # 检查主键的唯一性
            if not check_primary_key_uniqueness(df1, primary_keys) or not check_primary_key_uniqueness(df2,
                                                                                                       primary_keys):
                st.error("选择的主键在表格中不唯一，请重新选择主键。")
            else:
                result_df, df1_only, df2_only = compare_excels(file1, file2, primary_keys, selected_columns)

                # 显示对比概要信息
                total_rows = len(result_df)
                inconsistent_rows = len(result_df[result_df['比对结果'] == 'N'])
                st.write(f"共 {total_rows} 行，不一致的 {inconsistent_rows} 行。")

                # 显示对比结果
                st.write("共同数据对比结果：")
                st.write(result_df.to_html(escape=False), unsafe_allow_html=True)

                st.markdown("---")

                # 显示第一个表格有第二个表格没有的数据
                st.write("第一个表格有第二个表格没有的数据：")
                st.dataframe(df1_only)

                st.markdown("---")

                # 显示第二个表格有第一个表格没有的数据
                st.write("第二个表格有第一个表格没有的数据：")
                st.dataframe(df2_only)

                st.markdown("---")

                # 创建带有颜色标记的 Excel 文件
                wb = Workbook()
                wb = highlight_excel(wb, result_df, primary_keys,
                                     [col for col in selected_columns if col not in primary_keys])
                wb = write_df_to_excel(wb, df1_only, '第一个表格独有的数据')
                wb = write_df_to_excel(wb, df2_only, '第二个表格独有的数据')

                # 保存 Excel 文件到内存
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                # 下载对比结果
                st.download_button(
                    label="下载对比结果",
                    data=output,
                    file_name="对比结果.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


if __name__ == '__main__':
    st.sidebar.header("Excel Compare")
    excel_compare()
